import openpyxl
wb = openpyxl.load_workbook('D:/Tax-ETS/docs/zone.xlsx')
sheet = wb.active

# Show first 25 district names from zone.xlsx
print("zone.xlsx districts:")
for r in range(2, 27):
    z1 = sheet.cell(r, 2).value
    z2 = sheet.cell(r, 3).value
    if z1:
        print(f"  Z1: '{z1}'")
    if z2:
        print(f"  Z2: '{z2}'")

print("\n--- CSV first 10 ---")
with open('D:/Tax-ETS/docs/Lao_Districts_List.csv', encoding='utf-8-sig') as f:
    for i, line in enumerate(f.readlines()[1:11]):
        parts = line.strip().split(',')
        if len(parts) >= 3:
            print(f"  '{parts[2]}'")