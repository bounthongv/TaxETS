import openpyxl

wb = openpyxl.load_workbook('D:/Tax-ETS/docs/zone.xlsx')
sheet = wb.active

print(f"Rows: {sheet.max_row}, Columns: {sheet.max_column}")
print()

# Header
header = [cell.value for cell in sheet[1]]
print("Header:", header)
print()

# First 15 data rows
for r in range(2, min(16, sheet.max_row + 1)):
    row_data = [cell.value for cell in sheet[r]]
    print(f"Row {r}: {row_data}")