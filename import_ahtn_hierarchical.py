import openpyxl
import mysql.connector
import re
import os

# Database Config
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'tax_ets'
}

FILE_PATH = r"D:\Tax-ETS\docs\For print AHTN 2017 + NOR + WTO + ATIGA + AC+ AI+AANZ+AJ+AK+APTA+LV LATEST 12.9.19 .xlsx"

def parse_excel():
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

    print("Clearing existing data for fresh import...")
    cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
    cursor.execute("TRUNCATE TABLE bm_customs_chapters")
    cursor.execute("TRUNCATE TABLE bm_customs_sections")
    cursor.execute("TRUNCATE TABLE bm_customs_tariff")
    cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
    conn.commit()

    print(f"Loading workbook: {FILE_PATH}...")
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True, read_only=True)
    sheet = wb.active

    current_section_id = None
    current_chapter_id = None
    row_count = 0
    
    # Track grouping state
    last_val_a = None
    
    # Process rows
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), 1):
        val_a = str(row[0]).strip() if row[0] is not None else ""
        val_b = str(row[1]).strip() if row[1] is not None else "" # Sub code
        val_c = str(row[2]).strip() if row[2] is not None else "" # Description Lao
        val_d = str(row[3]).strip() if row[3] is not None else "" # Description Eng
        val_e = str(row[4]).strip() if row[4] is not None else "" # Unit
        
        # 1. Detect SECTION
        if "SECTION" in val_d.upper() and len(val_d) < 20:
            section_code = val_d
            name_lo = ""
            # Peek next row for name
            next_row = next(sheet.iter_rows(min_row=row_idx+1, max_row=row_idx+1, values_only=True))
            name_lo = str(next_row[0]).strip() if next_row[0] else ""
            name_en = str(next_row[3]).strip() if next_row[3] else ""
            
            cursor.execute("INSERT INTO bm_customs_sections (section_code, name_lo, name_en, order_idx) VALUES (%s, %s, %s, %s)",
                           (section_code, name_lo, name_en, row_idx))
            current_section_id = cursor.lastrowid
            conn.commit()
            print(f"Imported {section_code}: {name_en}")
            continue

        # 2. Detect CHAPTER
        if "CHAPTER" in val_d.upper() and len(val_d) < 20:
            chapter_code = val_d.replace("Chapter ", "")
            # Peek next row for name
            next_row = next(sheet.iter_rows(min_row=row_idx+1, max_row=row_idx+1, values_only=True))
            name_lo = str(next_row[0]).strip() if next_row[0] else ""
            name_en = str(next_row[3]).strip() if next_row[3] else ""
            
            cursor.execute("INSERT INTO bm_customs_chapters (section_id, chapter_code, name_lo, name_en, order_idx) VALUES (%s, %s, %s, %s, %s)",
                           (current_section_id, chapter_code, name_lo, name_en, row_idx))
            current_chapter_id = cursor.lastrowid
            conn.commit()
            # print(f"  - Chapter {chapter_code}")
            continue

        # 3. Detect Item Header (AHTN 2017 CODES header rows)
        if "AHTN" in val_a and "CODES" in val_a:
            continue
            
        # 4. Detect TARIFF ITEM OR CATEGORY HEADER
        # Rule: If it has an HS Code (looks like 0101.21.00) or it's a sub-header (starts with -)
        
        is_hs_code = re.match(r'^\d{2}\.\d{2}$|^\d{4}\.\d{2}\.\d{2}$|^\d{8}$|^\d{4}\.\d{2}$', val_a.replace(".",""))
        is_sub_header = val_c.startswith("-") or val_d.startswith("-")
        
        if is_hs_code or is_sub_header:
            # Determine hierarchy level based on number of dashes
            level = 0
            match = re.match(r'^(-+)', val_d)
            if match:
                level = len(match.group(1))
            
            is_header = 1 if not is_hs_code else 0
            
            # Extract rates
            rates = []
            for i in range(5, 15): # Normal, MFN, ATIGA, ACFTA, AKFTA...
                val = row[i]
                if isinstance(val, (int, float)):
                    rates.append(str(val))
                elif val is None:
                    rates.append("0")
                else:
                    rates.append(str(val))
            
            # Map rates to named indices
            # rate_normal, rate_mfn, rate_atiga, rate_acfta, rate_akfta, rate_ajcep, rate_aanzfta, rate_aifta, rate_apta, rate_laoviet
            
            sql = """INSERT INTO bm_customs_tariff 
                     (chapter_id, hs_code, sub_code, description_lo, description_en, unit, 
                      rate_normal, rate_mfn, rate_atiga, rate_acfta, rate_akfta, rate_ajcep, 
                      rate_aanzfta, rate_aifta, rate_apta, rate_laoviet, is_header, level, row_idx) 
                     VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            
            params = [current_chapter_id, val_a, val_b, val_c, val_d, val_e] + rates + [is_header, level, row_idx]
            cursor.execute(sql, params)
            row_count += 1
            
            if row_count % 1000 == 0:
                conn.commit()
                print(f"  Processed {row_count} rows...")

    conn.commit()
    cursor.close()
    conn.close()
    print(f"Import Finished. Total tariff rows: {row_count}")

if __name__ == "__main__":
    parse_excel()
