import openpyxl

wb = openpyxl.load_workbook('D:/Tax-ETS/docs/zone.xlsx')
sheet = wb.active

# Get district-province mapping from CSV
provinces = {}
with open('D:/Tax-ETS/docs/Lao_Districts_List.csv', encoding='utf-8-sig') as f:
    for line in f.readlines()[1:]:
        parts = line.strip().split(',')
        if len(parts) >= 3:
            d_name = parts[2].strip().lower()
            p_name = parts[4].strip().lower()
            provinces[d_name] = p_name

print(f"CSV districts: {len(provinces)}")

current_province = None
zone1_list = []
zone2_list = []
not_found = []

for r in range(2, sheet.max_row + 1):
    province = sheet.cell(r, 1).value
    z1 = sheet.cell(r, 2).value
    z2 = sheet.cell(r, 3).value
    
    if province:
        current_province = province.strip().lower()
    
    if z1:
        d = z1.strip().lower().replace(' district', '').replace(' district', '')
        if d in provinces:
            zone1_list.append(d)
        else:
            not_found.append(f"Z1: {z1}")
    
    if z2:
        d = z2.strip().lower().replace(' district', '').replace(' district', '')
        if d in provinces:
            zone2_list.append(d)
        else:
            not_found.append(f"Z2: {z2}")

print(f"\nZone 1: {len(zone1_list)}")
print(f"Zone 2: {len(zone2_list)}")
print(f"Not found: {len(not_found)}")

if not_found:
    print("\nFirst 20 not found:")
    for nf in not_found[:20]:
        print(f"  {nf}")

# Generate SQL
sql_file = 'D:/Tax-ETS/db/seed_zones.sql'
with open(sql_file, 'w', encoding='utf-8') as f:
    f.write("-- Zone import from docs/zone.xlsx\n")
    f.write("-- Generated: This updates districts with zone values\n\n")
    
    f.write("UPDATE districts SET zone = NULL;\n\n")
    
    f.write("-- Zone 1 (SEZ)\n")
    for d in zone1_list:
        f.write(f"UPDATE districts SET zone = 1 WHERE LOWER(district_name) = '{d}';\n")
    
    f.write("\n-- Zone 2 (Promotion)\n")
    for d in zone2_list:
        f.write(f"UPDATE districts SET zone = 2 WHERE LOWER(district_name) = '{d}';\n")

print(f"\nSQL: {sql_file}")